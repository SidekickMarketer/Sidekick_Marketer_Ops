#!/usr/bin/env python3
"""
Parse Social Data Script - Enhanced Version
Parses analytics files (CSV/XLSX) and post PDFs into unified JSON format
FIXED VERSION: Handles Meta 'Lifetime' date bug and GBP aggregate files
ENHANCED: Excel support, comprehensive statistics, better error reporting
"""

import argparse
import json
import csv
import re
import sys
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Optional
from collections import defaultdict


class SocialDataParser:
    """Parses social media data from various sources"""

    def __init__(self, verbose: bool = True):
        self.posts = []
        self.errors = []
        self.verbose = verbose
        self.stats = {
            'files_processed': 0,
            'files_failed': 0,
            'files_skipped': 0,
            'posts_parsed': 0,
            'posts_skipped': 0,
            'skip_reasons': defaultdict(int),
            'posts_by_platform': defaultdict(int),
            'date_formats_found': defaultdict(int)
        }

        # Try to import openpyxl for Excel support
        try:
            import openpyxl
            self.has_excel_support = True
        except ImportError:
            self.has_excel_support = False
            if verbose:
                print("⚠️  openpyxl not installed. Excel files will be skipped.")
                print("   Install with: pip install openpyxl")

    def parse_csv(self, file_path: Path) -> List[Dict[str, Any]]:
        """Parse CSV analytics file"""

        posts = []
        self.stats['files_processed'] += 1

        try:
            with open(file_path, 'r', encoding='utf-8-sig', errors='replace') as f:
                # Try to detect delimiter
                sample = f.read(1024)
                f.seek(0)

                if ',' in sample: delimiter = ','
                elif '\t' in sample: delimiter = '\t'
                elif ';' in sample: delimiter = ';'
                else: delimiter = ','

                reader = csv.DictReader(f, delimiter=delimiter)

                for row_num, row in enumerate(reader, start=2):
                    try:
                        # GBP Safety Check: Skip description rows
                        if self._is_description_row(row):
                            self.stats['posts_skipped'] += 1
                            self.stats['skip_reasons']['description_row'] += 1
                            continue

                        post = self._normalize_row(row, file_path.name)
                        if post:
                            posts.append(post)
                            self.stats['posts_parsed'] += 1
                            self.stats['posts_by_platform'][post['platform']] += 1
                        else:
                            self.stats['posts_skipped'] += 1
                            self.stats['skip_reasons']['normalization_failed'] += 1
                    except Exception as e:
                        # Log error but don't crash
                        self.errors.append(f"Row {row_num} in {file_path.name}: {str(e)}")
                        self.stats['posts_skipped'] += 1
                        self.stats['skip_reasons']['parse_error'] += 1

            if len(posts) > 0:
                if self.verbose:
                    print(f"✅ Parsed {len(posts)} posts from {file_path.name}")
            else:
                # Only warn if we expected posts but got none (silent skip for aggregates)
                if "gbp" not in file_path.name.lower():
                    if self.verbose:
                        print(f"⚠️  Parsed 0 posts from {file_path.name} (Check date formats)")
                    self.stats['files_skipped'] += 1

        except Exception as e:
            self.errors.append(f"Failed to parse {file_path.name}: {e}")
            self.stats['files_failed'] += 1

        return posts

    def parse_excel(self, file_path: Path) -> List[Dict[str, Any]]:
        """Parse Excel analytics file"""

        if not self.has_excel_support:
            self.stats['files_skipped'] += 1
            self.stats['skip_reasons']['no_excel_support'] += 1
            return []

        posts = []
        self.stats['files_processed'] += 1

        try:
            import openpyxl

            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

            # Try each sheet
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Get headers from first row
                headers = []
                for cell in sheet[1]:
                    if cell.value:
                        headers.append(str(cell.value).strip())
                    else:
                        headers.append('')

                if not headers:
                    continue

                # Parse data rows
                for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        # Create dict from headers and row values
                        row_dict = {}
                        for i, value in enumerate(row):
                            if i < len(headers) and headers[i]:
                                row_dict[headers[i]] = value

                        # Skip empty rows
                        if not any(row_dict.values()):
                            continue

                        # GBP Safety Check: Skip description rows
                        if self._is_description_row(row_dict):
                            self.stats['posts_skipped'] += 1
                            self.stats['skip_reasons']['description_row'] += 1
                            continue

                        post = self._normalize_row(row_dict, f"{file_path.name}:{sheet_name}")
                        if post:
                            posts.append(post)
                            self.stats['posts_parsed'] += 1
                            self.stats['posts_by_platform'][post['platform']] += 1
                        else:
                            self.stats['posts_skipped'] += 1
                            self.stats['skip_reasons']['normalization_failed'] += 1
                    except Exception as e:
                        self.errors.append(f"Row {row_num} in {file_path.name}:{sheet_name}: {str(e)}")
                        self.stats['posts_skipped'] += 1
                        self.stats['skip_reasons']['parse_error'] += 1

            if len(posts) > 0:
                if self.verbose:
                    print(f"✅ Parsed {len(posts)} posts from {file_path.name}")
            else:
                if self.verbose:
                    print(f"⚠️  Parsed 0 posts from {file_path.name}")
                self.stats['files_skipped'] += 1

        except Exception as e:
            self.errors.append(f"Failed to parse Excel {file_path.name}: {e}")
            self.stats['files_failed'] += 1

        return posts

    def _is_description_row(self, row: Dict[str, str]) -> bool:
        """Detects GBP secondary header rows that contain descriptions"""
        # If multiple values contain verbose descriptions, skip
        indicators = ['number of', 'total count', 'interactions with', 'people that viewed',
                     'times your', 'how many', 'accounts that']
        row_str = " ".join(str(v).lower() for v in row.values() if v)
        return sum(1 for ind in indicators if ind in row_str) >= 2

    def _normalize_row(self, row: Dict[str, Any], filename: str) -> Optional[Dict[str, Any]]:
        """Normalize a CSV/Excel row to standard format"""

        # Convert all values to strings for processing
        row = {k: str(v) if v is not None else '' for k, v in row.items()}

        # 1. Detect Platform
        platform = self._detect_platform(filename, row)

        # CRITICAL FIX: If this is a GBP Aggregate file (not posts), skip it
        if platform == 'gbp_aggregate':
            return None

        # 2. Extract Date (CRITICAL FIX APPLIED HERE)
        date = self._extract_date(row)
        if not date:
            self.stats['skip_reasons']['no_valid_date'] += 1
            return None

        # 3. Initialize Post Object
        post = {
            'date': date,
            'platform': platform,
            'source_file': filename
        }

        # 4. Mapping Logic
        mappings = {
            'post_type': ['type', 'post type', 'media_type', 'format', 'content type'],
            'caption': ['caption', 'description', 'text', 'post text', 'content', 'title'],
            'likes': ['likes', 'like count', 'reactions'],
            'comments': ['comments', 'comment count'],
            'shares': ['shares', 'share count'],
            'saves': ['saves', 'save count', 'saved', 'bookmarks'],
            'reach': ['reach', 'accounts reached', 'unique viewers'],
            'impressions': ['impressions', 'views', 'total views'],
            'link_clicks': ['link clicks', 'clicks', 'website clicks'],
            'engagement_rate': ['engagement rate', 'engagement'],
            'video_views': ['video views', 'plays', 'reel plays'],
            'permalink': ['permalink', 'post link', 'url']
        }

        row_lower = {k.lower().strip(): v for k, v in row.items() if v is not None and v != ''}

        for field, variations in mappings.items():
            for variation in variations:
                if variation in row_lower:
                    value = row_lower[variation]

                    # Clean numeric fields
                    if field in ['likes', 'comments', 'shares', 'saves', 'reach', 'impressions', 'link_clicks', 'video_views']:
                        post[field] = self._parse_number(value)
                    elif field == 'engagement_rate':
                        post[field] = self._parse_percentage(value)
                    else:
                        post[field] = value
                    break

        # Calculate engagement rate if missing
        if 'engagement_rate' not in post and post.get('reach', 0) > 0:
            total_eng = sum([post.get('likes',0), post.get('comments',0), post.get('shares',0), post.get('saves',0)])
            post['engagement_rate'] = round((total_eng / post['reach']) * 100, 2)

        # Calculate total engagement
        post['total_engagement'] = sum([post.get('likes',0), post.get('comments',0), post.get('shares',0), post.get('saves',0)])

        # Filter out empty rows (must have at least date + platform + one metric)
        return post if len(post) > 3 else None

    def _detect_platform(self, filename: str, row: Dict[str, str]) -> str:
        """Detect platform from filename or row content"""
        fn = filename.lower()

        # GBP Aggregate Check
        if 'gbp' in fn or 'google' in fn:
            # If row has headers like "Store code" or "Business name", it's aggregate, not posts
            if any(k in row for k in ['Store code', 'Business name', 'Total views']):
                return 'gbp_aggregate'
            return 'google_business_profile'

        if 'instagram' in fn or 'ig' in fn: return 'instagram'
        if 'facebook' in fn or 'fb' in fn: return 'facebook'
        if 'linkedin' in fn: return 'linkedin'
        if 'twitter' in fn or 'x.com' in fn: return 'twitter'
        if 'tiktok' in fn: return 'tiktok'
        if 'youtube' in fn: return 'youtube'

        return 'unknown'

    def _extract_date(self, row: Dict[str, str]) -> Optional[str]:
        """Extract date from row - FIXED PRIORITY"""

        # CRITICAL FIX: 'publish time' is FIRST because 'date' contains "Lifetime" garbage
        date_fields = ['publish time', 'publish date', 'posted date', 'posted', 'created', 'created time', 'timestamp', 'date']

        row_lower = {k.lower().strip(): v for k, v in row.items() if v}

        for field in date_fields:
            if field in row_lower:
                date_str = row_lower[field]
                # Skip Meta's "Lifetime" garbage
                if "lifetime" in str(date_str).lower():
                    continue

                parsed_date = self._parse_date(date_str)
                if parsed_date:
                    return parsed_date
        return None

    def _parse_date(self, date_str: str) -> Optional[str]:
        """Parse date string to YYYY-MM-DD format"""
        if not date_str or str(date_str).strip() == '': return None

        date_str = str(date_str).strip()

        # Skip non-date values
        if date_str.lower() in ['lifetime', 'n/a', 'null', 'none', '']:
            return None

        # Meta Export Format (MM/DD/YYYY HH:MM)
        formats = [
            '%m/%d/%Y %H:%M',    # Meta default (03/25/2025 22:00)
            '%Y-%m-%d',
            '%m/%d/%Y',
            '%d/%m/%Y',
            '%Y/%m/%d',
            '%Y-%m-%d %H:%M:%S',
            '%m/%d/%Y %H:%M:%S',
            '%d/%m/%Y %H:%M:%S',
            '%Y-%m-%d %H:%M',
            '%B %d, %Y',         # January 15, 2025
            '%b %d, %Y',         # Jan 15, 2025
            '%d %B %Y',          # 15 January 2025
            '%d %b %Y',          # 15 Jan 2025
        ]

        for fmt in formats:
            try:
                dt = datetime.strptime(date_str, fmt)
                self.stats['date_formats_found'][fmt] += 1
                return dt.strftime('%Y-%m-%d')
            except ValueError:
                continue

        # Regex fallback for YYYY-MM-DD
        match = re.search(r'(\d{4})-(\d{2})-(\d{2})', date_str)
        if match:
            self.stats['date_formats_found']['regex_yyyy-mm-dd'] += 1
            return match.group(0)

        return None

    def _parse_number(self, value: str) -> int:
        """Parse number from string with robust error handling"""
        if not value: return 0

        # Cleanup string
        clean = str(value).lower().strip().replace(',', '').replace(' ', '')
        if clean == '': return 0

        # Handle K/M notation
        multiplier = 1
        if 'k' in clean:
            multiplier = 1000
            clean = clean.replace('k', '')
        elif 'm' in clean:
            multiplier = 1000000
            clean = clean.replace('m', '')

        try:
            return int(float(clean) * multiplier)
        except:
            return 0

    def _parse_percentage(self, value: str) -> float:
        if not value: return 0.0
        clean = str(value).replace('%', '').strip()
        try: return float(clean)
        except: return 0.0

    def search_directory(self, directory: Path, recursive: bool = True) -> List[Dict[str, Any]]:
        """Search directory for analytics files"""
        all_posts = []

        # Find CSV files
        csv_pattern = "**/*.csv" if recursive else "*.csv"
        csv_files = list(directory.glob(csv_pattern))

        # Find Excel files
        xlsx_pattern = "**/*.xlsx" if recursive else "*.xlsx"
        xls_pattern = "**/*.xls" if recursive else "*.xls"
        excel_files = list(directory.glob(xlsx_pattern)) + list(directory.glob(xls_pattern))

        total_files = len(csv_files) + len(excel_files)

        if self.verbose:
            print(f"\n📂 Searching {directory.name}...")
            print(f"   Found {len(csv_files)} CSV files")
            print(f"   Found {len(excel_files)} Excel files")
            print(f"   Total: {total_files} files\n")

        # Parse CSV files
        for f in csv_files:
            all_posts.extend(self.parse_csv(f))

        # Parse Excel files
        for f in excel_files:
            all_posts.extend(self.parse_excel(f))

        return all_posts

    def save_json(self, posts: List[Dict[str, Any]], output_path: Path):
        """Save parsed posts to JSON with metadata"""
        # Dedup logic based on date+platform+caption(partial)
        unique_posts = []
        seen = set()

        # Sort by date
        posts.sort(key=lambda x: x['date'])

        for p in posts:
            # Create a unique key
            key = f"{p['date']}_{p['platform']}_{str(p.get('likes',0))}_{str(p.get('caption',''))[:50]}"
            if key not in seen:
                seen.add(key)
                unique_posts.append(p)

        # Calculate platform breakdown
        platform_breakdown = defaultdict(int)
        for p in unique_posts:
            platform_breakdown[p['platform']] += 1

        output = {
            'metadata': {
                'generated_at': datetime.now().isoformat(),
                'total_posts': len(unique_posts),
                'duplicates_removed': len(posts) - len(unique_posts),
                'date_range': {
                    'start': unique_posts[0]['date'] if unique_posts else None,
                    'end': unique_posts[-1]['date'] if unique_posts else None
                },
                'platforms': dict(platform_breakdown),
                'parsing_stats': {
                    'files_processed': self.stats['files_processed'],
                    'files_failed': self.stats['files_failed'],
                    'files_skipped': self.stats['files_skipped'],
                    'posts_parsed': self.stats['posts_parsed'],
                    'posts_skipped': self.stats['posts_skipped'],
                    'errors': len(self.errors)
                }
            },
            'posts': unique_posts
        }

        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(output, f, indent=2)

        if self.verbose:
            print(f"\n💾 Saved {len(unique_posts)} posts to {output_path}")

    def print_statistics(self):
        """Print comprehensive parsing statistics"""
        print(f"\n{'='*70}")
        print(f"PARSING STATISTICS")
        print(f"{'='*70}")
        print(f"\n📊 FILES:")
        print(f"   Processed: {self.stats['files_processed']}")
        print(f"   Failed:    {self.stats['files_failed']}")
        print(f"   Skipped:   {self.stats['files_skipped']}")

        print(f"\n📝 POSTS:")
        print(f"   Parsed:  {self.stats['posts_parsed']}")
        print(f"   Skipped: {self.stats['posts_skipped']}")

        if self.stats['posts_by_platform']:
            print(f"\n📱 BY PLATFORM:")
            for platform, count in sorted(self.stats['posts_by_platform'].items()):
                print(f"   {platform.title()}: {count}")

        if self.stats['skip_reasons']:
            print(f"\n⚠️  SKIP REASONS:")
            for reason, count in sorted(self.stats['skip_reasons'].items(), key=lambda x: x[1], reverse=True):
                print(f"   {reason}: {count}")

        if self.stats['date_formats_found']:
            print(f"\n📅 DATE FORMATS DETECTED:")
            for fmt, count in sorted(self.stats['date_formats_found'].items(), key=lambda x: x[1], reverse=True)[:5]:
                print(f"   {fmt}: {count} times")

        if self.errors:
            print(f"\n❌ ERRORS ({len(self.errors)} total):")
            for error in self.errors[:10]:  # Show first 10
                print(f"   {error}")
            if len(self.errors) > 10:
                print(f"   ... and {len(self.errors) - 10} more")

        print(f"\n{'='*70}\n")

def main():
    parser = argparse.ArgumentParser(
        description='Parse social media analytics files (CSV/Excel) into unified JSON format',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Parse specific files
  %(prog)s --analytics file1.csv file2.xlsx --output results.json

  # Search directory recursively
  %(prog)s --search-dir ./social_data --output results.json

  # Search directory non-recursively
  %(prog)s --search-dir ./social_data --no-recursive --output results.json

  # Quiet mode
  %(prog)s --search-dir ./social_data --output results.json --quiet
        """
    )
    parser.add_argument('--analytics', nargs='+', help='Specific analytics files to parse')
    parser.add_argument('--search-dir', help='Directory to search for analytics files')
    parser.add_argument('--output', required=True, help='Output JSON file path')
    parser.add_argument('--no-recursive', action='store_true', help='Disable recursive directory search')
    parser.add_argument('--quiet', action='store_true', help='Suppress progress messages')

    args = parser.parse_args()

    verbose = not args.quiet
    parser_obj = SocialDataParser(verbose=verbose)
    all_posts = []

    if args.analytics:
        for f in args.analytics:
            file_path = Path(f)
            if not file_path.exists():
                print(f"❌ File not found: {f}")
                continue

            if file_path.suffix.lower() == '.csv':
                all_posts.extend(parser_obj.parse_csv(file_path))
            elif file_path.suffix.lower() in ['.xlsx', '.xls']:
                all_posts.extend(parser_obj.parse_excel(file_path))
            else:
                print(f"⚠️  Unsupported file type: {f}")

    if args.search_dir:
        path = Path(args.search_dir)
        if path.exists():
            recursive = not args.no_recursive
            all_posts.extend(parser_obj.search_directory(path, recursive=recursive))
        else:
            print(f"❌ Directory not found: {args.search_dir}")

    if all_posts:
        parser_obj.save_json(all_posts, Path(args.output))
        if verbose:
            parser_obj.print_statistics()
    else:
        print("❌ No posts found. Check file paths and date formats.")
        if verbose:
            parser_obj.print_statistics()
        sys.exit(1)

if __name__ == "__main__":
    main()
